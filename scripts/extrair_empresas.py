#!/usr/bin/env python3
"""
Script para extrair e analisar empresas
1. Ler empresas.xlsx 
2. Comparar com coluna '## Empresa' do CSV de transações
3. Gerar b_dados_empresas.csv estruturado
"""

import csv
import re
from pathlib import Path
from collections import defaultdict

def extrair_empresas_transacoes():
    """Extrai empresas únicas da coluna ## Empresa do CSV de transações"""
    base_path = Path(__file__).parent.parent
    transacoes_file = base_path / 'importacao' / 'b_dados_transacoes.csv'
    
    empresas_transacoes = set()
    
    try:
        with open(transacoes_file, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            
            for row in reader:
                empresa = row.get('## Empresa', '').strip()
                if empresa:
                    empresas_transacoes.add(empresa)
    
    except Exception as e:
        print(f"Erro ao ler transações: {e}")
        return set()
    
    return empresas_transacoes

def tentar_ler_xlsx():
    """Tenta diferentes métodos para ler o arquivo Excel"""
    base_path = Path(__file__).parent.parent
    excel_file = base_path / 'informacoes_llm_visual' / 'referencia_empresas' / 'empresas.xlsx'
    
    print(f"Tentando ler: {excel_file}")
    
    # Verificar se arquivo existe
    if not excel_file.exists():
        print(f"❌ Arquivo não encontrado: {excel_file}")
        return []
    
    # Método 1: Tentar com pandas (se disponível)
    try:
        import pandas as pd
        df = pd.read_excel(excel_file)
        print("✅ Arquivo Excel lido com pandas")
        return df.to_dict('records')
    except ImportError:
        print("⚠️ Pandas não disponível")
    except Exception as e:
        print(f"❌ Erro com pandas: {e}")
    
    # Método 2: Tentar com openpyxl (se disponível)
    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_file)
        ws = wb.active
        
        # Ler cabeçalho
        headers = [cell.value for cell in ws[1]]
        
        # Ler dados
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):  # Se linha não está vazia
                row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
                data.append(row_dict)
        
        print("✅ Arquivo Excel lido com openpyxl")
        return data
        
    except ImportError:
        print("⚠️ openpyxl não disponível")
    except Exception as e:
        print(f"❌ Erro com openpyxl: {e}")
    
    print("❌ Não foi possível ler o arquivo Excel")
    return []

def normalizar_nome_empresa(nome):
    """Normaliza nome da empresa para comparação"""
    if not nome:
        return ""
    
    # Remover números e códigos do início
    nome = re.sub(r'^\d+\s*-\s*', '', str(nome))
    
    # Limpar e normalizar
    nome = nome.strip().upper()
    
    return nome

def extrair_numero_empresa(texto):
    """Extrai o número da empresa do formato '0001 - NOME'"""
    if not texto:
        return None
    
    match = re.match(r'^(\d+)\s*-', str(texto))
    return match.group(1) if match else None

def gerar_cnpj_placeholder(numero):
    """Gera um CNPJ placeholder baseado no número da empresa"""
    if not numero:
        return ""
    
    # Formato: XX.XXX.XXX/0001-00 onde XX.XXX.XXX é baseado no número
    numero_str = str(numero).zfill(8)
    return f"{numero_str[:2]}.{numero_str[2:5]}.{numero_str[5:8]}/0001-00"

def main():
    """Executa o processo de extração"""
    print("="*60)
    print("EXTRAÇÃO DE EMPRESAS")
    print("="*60)
    
    # 1. Extrair empresas das transações
    print("\n1. Extraindo empresas das transações...")
    empresas_transacoes = extrair_empresas_transacoes()
    print(f"   Encontradas {len(empresas_transacoes)} empresas únicas")
    
    for empresa in sorted(empresas_transacoes):
        print(f"   - {empresa}")
    
    # 2. Tentar ler arquivo Excel
    print("\n2. Tentando ler arquivo Excel...")
    dados_excel = tentar_ler_xlsx()
    
    if dados_excel:
        print(f"   Encontradas {len(dados_excel)} empresas no Excel")
        print("   Colunas:", list(dados_excel[0].keys()) if dados_excel else "Nenhuma")
    
    # 3. Criar dataset combinado
    print("\n3. Criando dataset combinado...")
    
    empresas_finais = []
    id_counter = 1
    
    # Processar empresas das transações
    for empresa_transacao in sorted(empresas_transacoes):
        numero = extrair_numero_empresa(empresa_transacao)
        nome_limpo = normalizar_nome_empresa(empresa_transacao)
        
        # Buscar dados correspondentes no Excel
        dados_excel_match = None
        if dados_excel:
            for item in dados_excel:
                # Tentar diferentes campos para comparação
                campos_nome = ['nome', 'razao_social', 'empresa', 'Nome', 'Razão Social', 'Empresa']
                for campo in campos_nome:
                    if campo in item and item[campo]:
                        if normalizar_nome_empresa(item[campo]) == nome_limpo:
                            dados_excel_match = item
                            break
                if dados_excel_match:
                    break
        
        # Montar registro
        empresa_final = {
            'id': id_counter,
            'codigo': numero or str(id_counter).zfill(4),
            'nome': nome_limpo or 'EMPRESA SEM NOME',
            'nome_original': empresa_transacao,
            'grupo': 'SELLETA',  # Assumir que todas fazem parte do grupo Selleta
            'cnpj': '',
            'endereco': '',
            'municipio': '',
            'ativo': 1
        }
        
        # Complementar com dados do Excel se encontrados
        if dados_excel_match:
            # Mapear campos do Excel para nossa estrutura
            campos_mapeamento = {
                'cnpj': ['cnpj', 'CNPJ', 'documento'],
                'endereco': ['endereco', 'endereço', 'Endereço', 'address'],
                'municipio': ['municipio', 'município', 'cidade', 'Município', 'Cidade'],
                'grupo': ['grupo', 'Grupo', 'holding']
            }
            
            for campo_nosso, campos_excel in campos_mapeamento.items():
                for campo_excel in campos_excel:
                    if campo_excel in dados_excel_match and dados_excel_match[campo_excel]:
                        empresa_final[campo_nosso] = str(dados_excel_match[campo_excel]).strip()
                        break
        
        # Se não tem CNPJ, gerar placeholder
        if not empresa_final['cnpj']:
            empresa_final['cnpj'] = gerar_cnpj_placeholder(numero or id_counter)
        
        empresas_finais.append(empresa_final)
        id_counter += 1
    
    # 4. Salvar CSV
    print(f"\n4. Salvando {len(empresas_finais)} empresas...")
    
    base_path = Path(__file__).parent.parent
    output_file = base_path / 'importacao' / 'b_dados_empresas.csv'
    
    with open(output_file, 'w', encoding='utf-8-sig', newline='') as f:
        fieldnames = ['id', 'codigo', 'nome', 'grupo', 'cnpj', 'endereco', 'municipio', 'ativo']
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        
        for empresa in empresas_finais:
            # Filtrar apenas campos necessários para o CSV final
            row = {campo: empresa.get(campo, '') for campo in fieldnames}
            writer.writerow(row)
    
    print(f"✅ Arquivo salvo: {output_file}")
    
    # 5. Estatísticas
    print("\n📊 ESTATÍSTICAS:")
    print(f"   Total de empresas: {len(empresas_finais)}")
    
    grupos = defaultdict(int)
    municipios = defaultdict(int)
    
    for empresa in empresas_finais:
        grupos[empresa.get('grupo', 'SEM GRUPO')] += 1
        municipio = empresa.get('municipio', 'SEM MUNICÍPIO')
        if municipio:
            municipios[municipio] += 1
    
    print(f"   Grupos: {dict(grupos)}")
    print(f"   Municípios: {dict(municipios)}")
    
    # 6. Preview
    print("\n📋 PREVIEW (primeiras 5 empresas):")
    for i, empresa in enumerate(empresas_finais[:5]):
        print(f"   {i+1}. {empresa['codigo']} - {empresa['nome']}")
        print(f"      CNPJ: {empresa['cnpj']}")
        print(f"      Município: {empresa['municipio'] or 'Não informado'}")
        print()

if __name__ == "__main__":
    main()